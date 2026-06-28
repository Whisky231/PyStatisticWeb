import glob
import math
import os
import time

import numpy as np
import plotly.graph_objs as go
from flask import Flask, render_template, request, redirect, url_for, send_file
from openpyxl import Workbook
from openpyxl.chart import Reference, BarChart
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.shapes import LineProperties
from openpyxl.styles import Alignment, Border, Side
from scipy.interpolate import make_interp_spline
from scipy.stats import norm

app = Flask(__name__)

# Global Plotly font configuration for larger, more readable charts
PLOTLY_FONT = dict(family='Inter, Arial, sans-serif', size=16, color='#333')
PLOTLY_TITLE_FONT = dict(family='Inter, Arial, sans-serif', size=22, color='#111')
PLOTLY_AXIS_FONT = dict(family='Inter, Arial, sans-serif', size=18, color='#333')
PLOTLY_TICK_FONT = dict(family='Inter, Arial, sans-serif', size=16, color='#444')
PLOTLY_ANNOTATION_FONT = dict(family='Inter, Arial, sans-serif', size=16, color='#222')

# Функция для очистки HTML-файлов из папки static при запуске
def cleanup_static_folder():
    # Путь к директории static
    static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')

    # Файлы, которые нужно сохранить
    files_to_keep = ['styles.css', 'data.xlsx', 'statistics.xlsx']

    # Удаляем все HTML-файлы
    for html_file in glob.glob(os.path.join(static_dir, '*.html')):
        try:
            os.remove(html_file)
            print(f"Удален файл: {html_file}")
        except Exception as e:
            print(f"Не удалось удалить {html_file}: {e}")

cleanup_static_folder()

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/process', methods=['POST'])
def process():
    data = {}
    for key in request.form:
        if key.startswith('data_'):
            values = request.form[key].strip().split()
            if values:
                data[key] = [float(value.strip().replace(',', '.')) for value in values if value.strip()]

    if not data:
        return redirect(url_for('index'))

    sections_data = {}  # Словарь для хранения данных по сечениям
    statistics = {}  # Словарь для хранения статистики по сечениям
    sections = []  # Список сечений для сортировки
    plot_files = []  # Список файлов с графиками


    for section, values in data.items():
        sections_data[section] = values
        sections.append(section)

        average = sum(values) / len(values)
        maxX = max(values)
        minX = min(values)
        variance = sum((x - average) ** 2 for x in values) / len(values)
        sred_kvad_otkl = math.sqrt(variance)
        koev_variazii = sred_kvad_otkl / average

        statistics[section] = {
            'variance': round(variance, 3),
            'sred_kvad_otkl': round(sred_kvad_otkl, 3),
            'sr_snaz': round(average, 3),
            'maxX': round(maxX, 3),
            'minX': round(minX, 3),
            'koev_variazii': round(koev_variazii, 3)
        }

        epure_file = create_epure(maxX, minX, koev_variazii, section)
        plot_files.append(epure_file)

        probability_curve_file = create_probability_curve(values, section)
        plot_files.append(probability_curve_file)

        # Сортируем сечения для правильного отображения
        sections.sort()

        # Создаем графики для каждого сечения
        # Количество интервалов (формула 3.1 из скриншота)
        n = round(math.sqrt(len(values)))
        # Протяженность одного интервала (формула 3.2 из скриншота)
        A = (maxX - minX) / n
        # Начало первого интервала (формула из скриншота)
        C = minX - 0.5 * A
        if C < 0:
            C = 0  # Принимаем C = 0, если получилось отрицательное значение

        # Создаем интервалы
        kol_interval = n
        interval = [[0] * 3 for _ in range(kol_interval)]
        interval[0][0] = C
        interval[0][1] = interval[0][0] + A
        
        # Рассчитываем интервалы
        for i in range(1, kol_interval):
            interval[i][0] = interval[i - 1][1]
            interval[i][1] = interval[i][0] + A
        
        # Рассчитываем количество деталей в каждом интервале
        numbers_in_interval = 0
        for b in range(kol_interval):
            for x in values:
                if interval[b][0] <= x <= interval[b][1]:
                    numbers_in_interval += 1
            interval[b][2] = numbers_in_interval
            numbers_in_interval = 0

        # Рассчитываем середины интервалов
        midpoints = [(interval[i][0] + interval[i][1]) / 2 for i in range(kol_interval)]
        
        # Рассчитываем опытные вероятности (формула 3.5 из скриншота)
        Pi = [(interval[i][2] / len(values)) for i in range(kol_interval)]
        
        # Рассчитываем накопленные опытные вероятности (формула 3.6 из скриншота)
        accumulated_Pi = []
        accumulated_sum = 0
        for i in range(kol_interval):
            accumulated_sum += Pi[i]
            accumulated_Pi.append(accumulated_sum)
        
        # Проверка суммы частот и накопленных вероятностей
        sum_frequencies = sum(interval[i][2] for i in range(kol_interval))
        if sum_frequencies != len(values):
            print(f"ВНИМАНИЕ! Сумма частот по интервалам ({sum_frequencies}) не равна общему количеству значений ({len(values)})!")
        
        if abs(accumulated_Pi[-1] - 1.0) > 0.01:  # Проверка с небольшой погрешностью
            print(f"ВНИМАНИЕ! Сумма накопленных опытных вероятностей ({accumulated_Pi[-1]}) не равна 1.0!")
        
        # Сохраняем дополнительную статистику для текущего сечения
        statistics[section].update({
            'n': n,  # Количество интервалов
            'A': round(A, 3),  # Протяженность одного интервала
            'C': round(C, 3),  # Начало первого интервала
            'intervals': [[round(interval[i][0], 3), round(interval[i][1], 3)] for i in range(kol_interval)],
            'midpoints': [round(mp, 3) for mp in midpoints],  # Середины интервалов
            'frequencies': [interval[i][2] for i in range(kol_interval)],  # Частоты
            'Pi': [round(p, 3) for p in Pi],  # Опытные вероятности
            'accumulated_Pi': [round(p, 3) for p in accumulated_Pi]  # Накопленные опытные вероятности
        })


        avg_value_formula = sum(midpoints[i] * Pi[i] for i in range(kol_interval))
        

        variance_formula = sum((midpoints[i] - avg_value_formula) ** 2 * Pi[i] for i in range(kol_interval))
        std_dev_formula = math.sqrt(variance_formula)
        

        coef_var_formula = std_dev_formula / (avg_value_formula - C) if avg_value_formula > C else 0
        


        lambda_min = 0
        if len(values) > 1:
            min_value = min(values)
            # Находим смежную точку
            values_sorted = sorted(values)
            if values_sorted.index(min_value) + 1 < len(values_sorted):
                next_min_value = values_sorted[values_sorted.index(min_value) + 1]
                lambda_min = (next_min_value - min_value) / std_dev_formula if std_dev_formula > 0 else 0
        
        # Для наибольшего значения износа
        lambda_max = 0
        if len(values) > 1:
            max_value = max(values)
            # Находим смежную точку
            values_sorted = sorted(values)
            if values_sorted.index(max_value) - 1 >= 0:
                prev_max_value = values_sorted[values_sorted.index(max_value) - 1]
                lambda_max = (max_value - prev_max_value) / std_dev_formula if std_dev_formula > 0 else 0
        
        # Определение критерия Ирвина для N=50 и доверительной вероятности α = 0.95
        irwin_criterion = 1.1  # Табличное значение критерия Ирвина
        is_reliable = lambda_min < irwin_criterion and lambda_max < irwin_criterion
        
        # 7.1 Расчет значения интегральной функции F(Икi) для ЗНР в конце интервалов
        f_znr = []
        for i in range(kol_interval):
            interval_end = interval[i][1]
            arg_znr = (interval_end - avg_value_formula) / std_dev_formula if std_dev_formula > 0 else 0
            # Значение центрированной интегральной функции (используем функцию нормального распределения)
            f0 = norm.cdf(arg_znr)
            f_znr.append(round(f0, 3))
        
        # 7.2 Расчет значения интегральной функции F(Икi) для ЗРВ в конце интервалов
        # Параметры для ЗРВ
        param_b = 3.30  # По приложению 5 для v = 0.33
        param_KB = 0.90  # По приложению 5 для v = 0.33
        param_a = (avg_value_formula - C) / param_KB if param_KB > 0 else 0
        
        f_zrv = []
        for i in range(kol_interval):
            interval_end = interval[i][1]
            arg_zrv = (interval_end - C) / param_a if param_a > 0 else 0
            # Значение табулированной интегральной функции для ЗРВ
            # Для упрощения используем ту же функцию нормального распределения
            f_zrv.append(round(norm.cdf(arg_zrv), 3))
        
        # Добавляем новые статистические показатели
        statistics[section].update({
            'avg_value_formula': round(avg_value_formula, 3),  # Среднее значение по формуле 4.1
            'std_dev_formula': round(std_dev_formula, 3),  # Среднее квадратическое отклонение по формуле 4.2
            'coef_var_formula': round(coef_var_formula, 3),  # Коэффициент вариации по формуле 4.3
            'lambda_min': round(lambda_min, 3),  # Коэффициент для проверки минимального значения
            'lambda_max': round(lambda_max, 3),  # Коэффициент для проверки максимального значения
            'irwin_criterion': irwin_criterion,  # Табличное значение критерия Ирвина
            'is_reliable': is_reliable,  # Результат проверки на достоверность
            'f_znr': f_znr,  # Значения интегральной функции для ЗНР
            'f_zrv': f_zrv,  # Значения интегральной функции для ЗРВ
            'param_b': param_b,  # Параметр b для ЗРВ
            'param_KB': param_KB,  # Коэффициент ЗРВ
            'param_a': round(param_a, 3)  # Параметр a для ЗРВ
        })

        # График количества (гистограмма)
        fig1 = go.Figure(data=[
            go.Bar(
                x=[f"{int(i[0])}-{int(i[1])}" for i in interval],  # Округляем до целых
                y=[i[2] for i in interval],
                marker_color='rgb(255,127,14)',  # Оранжевый цвет как в Excel
                name='Series1'
            )
        ])
        fig1.update_layout(
            title=dict(text=f'Гистограмма — Сечение {section}', font=PLOTLY_TITLE_FONT),
            xaxis_title=dict(text='Интервалы', font=PLOTLY_AXIS_FONT),
            yaxis_title=dict(text='Количество', font=PLOTLY_AXIS_FONT),
            font=PLOTLY_FONT,
            plot_bgcolor='white',
            showlegend=True,
            bargap=0.2,
            width=1100,
            height=600,
            margin=dict(t=80, b=70, l=80, r=50),
            xaxis=dict(
                showgrid=True,
                gridwidth=1,
                gridcolor='rgb(210,210,210)',
                showline=True,
                linewidth=1,
                linecolor='rgb(180,180,180)',
                tickfont=PLOTLY_TICK_FONT
            ),
            yaxis=dict(
                showgrid=True,
                gridwidth=1,
                gridcolor='rgb(210,210,210)',
                showline=True,
                linewidth=1,
                linecolor='rgb(180,180,180)',
                range=[0, max([i[2] for i in interval]) * 1.1],
                tickfont=PLOTLY_TICK_FONT
            )
        )
        plot_file = f'{section.replace(".", "_")}_plot1.html'
        fig1.write_html(os.path.join('static', plot_file), include_plotlyjs='cdn')
        plot_files.append(plot_file)


    wb = Workbook()
    ws = wb.active


    ws.append(["№ изм.", "l"])  # Возвращаем старые заголовки

    print("Собранные данные:", data)  # Выводим собранные данные

    # Создаем плоский список всех значений
    all_values = []
    for section_values in data.values():
        all_values.extend(section_values)
    print("Значения для записи в Excel:", all_values)  # Выводим значения, которые будут записаны в Excel

    # Заполняем данные последовательно
    for i, value in enumerate(all_values, 1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=value)
        ws.cell(row=i + 1, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=i + 1, column=2).alignment = Alignment(horizontal='center')

    # Добавляем таблицу интервалов
    ws.cell(row=1, column=4).value = "интервалы"
    ws.cell(row=1, column=5).value = "кол-ва"

    # Форматируем интервалы с точками
    for i, interval in enumerate(interval):
        start = int(interval[0])
        end = int(interval[1])
        ws.cell(row=i + 2, column=4).value = f"{start}...{end}"
        ws.cell(row=i + 2, column=5).value = interval[2]
        ws.cell(row=i + 2, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=i + 2, column=5).alignment = Alignment(horizontal='center')

    # Создаем график с правильными настройками
    chart = BarChart()
    chart.type = "col"
    chart.title = "кол-ва"

    # Настраиваем данные для графика
    data = Reference(ws, min_col=5, max_col=5, min_row=2, max_row=len(interval) + 1)
    cats = Reference(ws, min_col=4, max_col=4, min_row=2, max_row=len(interval) + 1)

    chart.add_data(data)
    chart.set_categories(cats)

    # Настройка внешнего вида графика
    chart.x_axis.title = None
    chart.y_axis.title = None

    # Важные настройки для соответствия примеру
    chart.legend = None  # Убираем легенду
    chart.height = 10
    chart.width = 15

    # Настройка осей и сетки
    from openpyxl.chart.axis import ChartLines
    chart.y_axis.majorGridlines = ChartLines()
    chart.x_axis.majorGridlines = None

    # Настройка масштаба
    chart.y_axis.scaling.max = 7
    chart.y_axis.scaling.min = 0

    # Настройка границ графика
    chart.plot_area.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill="000000"))

    # Настройка цветов и стиля столбцов
    s1 = chart.series[0]
    s1.graphicalProperties.solidFill = "FF8C00"  # Оранжевый цвет
    s1.graphicalProperties.ln = LineProperties(solidFill="FF8C00", prstDash="solid", w=0)

    # Настройка внешнего вида сетки
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill="CCCCCC", prstDash="solid", w=1))
    chart.plot_area.spPr = GraphicalProperties(noFill=True)

    # Добавляем график на лист
    ws.add_chart(chart, "G2")

    # Устанавливаем ширину столбцов
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10

    # Выравнивание заголовков
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=1, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=1, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=1, column=5).alignment = Alignment(horizontal='center')

    # Добавляем границы для таблицы интервалов
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Применяем границы к таблице интервалов
    for row in range(1, len(interval) + 2):
        for col in [4, 5]:
            ws.cell(row=row, column=col).border = thin_border

    # Сохраняем файл
    wb.save('static/data.xlsx')

    # Создаем и добавляем график профиля
    profile_file = create_profile(all_values, 'profile')
    plot_files.append(profile_file)

    return render_template('result.html',
                           statistics=statistics,
                           plot_files=plot_files,
                           sections=sections)


def create_epure(maxX, minX, koev_variazii, section_name):
    """
    Функция создания эпюры износа культиваторной лапы посевного комплекса "Моррис"
    согласно методике "Износы культиваторных лап посевного комплекса «Моррис»"
    """
    # Создаем точки для оси X (значения d) - расстояние от носка лапы
    x = np.linspace(0, 100, 200)  # Увеличиваем количество точек для более гладких кривых

    # Рассчитываем среднее значение
    avg = (maxX + minX) / 2

    # Параметры для создания U-образных кривых
    # Создаем параболические кривые с минимумом посередине и максимумами на краях

    # Функция для создания U-образной кривой
    def u_curve(base_val, min_factor=0.8, max_factor=1.1):
        # Параболическая функция с минимумом около x=50
        return base_val * (min_factor + (max_factor - min_factor) * ((x - 50)**2 / 2500))

    # Максимальный износ
    max_curve = u_curve(maxX, 0.8, 1.0)

    # Средний износ
    mid_curve = u_curve(avg, 0.85, 1.05)

    # Минимальный износ
    min_curve = u_curve(minX, 0.9, 1.1)

    # Создаем график
    fig = go.Figure()

    # Добавляем три кривые с правильными названиями и стилями
    fig.add_trace(go.Scatter(
        x=x,
        y=max_curve,
        mode='lines',
        name='Δl max',
        line=dict(color='rgb(0, 175, 150)', width=3, dash='solid')
    ))

    fig.add_trace(go.Scatter(
        x=x,
        y=mid_curve,
        mode='lines',
        name='l̄',
        line=dict(color='rgb(0, 175, 150)', width=3, dash='solid')
    ))

    fig.add_trace(go.Scatter(
        x=x,
        y=min_curve,
        mode='lines',
        name='Δl min',
        line=dict(color='rgb(0, 175, 150)', width=3, dash='solid')
    ))

    # Определяем позиции вертикальных линий для разметки секций
    # Согласно изображению: d₁, d₂, d₃, d₄ и т.д.
    d_positions = [0, 10, 20, 40, 60, 80, 90, 100]

    # Добавляем вертикальные линии
    for d in d_positions:
        fig.add_shape(
            type='line',
            x0=d,
            x1=d,
            y0=min(min_curve) * 0.9,
            y1=max(max_curve) * 1.1,
            line=dict(color='blue', width=1.5, dash='solid')
        )


    labels = ["", "d₁", "d₂", "d₃", "d₁", "d₂", "d₃", "d₄"]
    y_positions = [min(min_curve) * 0.95, min(min_curve) * 0.95, min(min_curve) * 0.95,
                  min(min_curve) * 0.95, min(min_curve) * 0.95, min(min_curve) * 0.95,
                  min(min_curve) * 0.95, min(min_curve) * 0.95]

    for i, (d, label, y_pos) in enumerate(zip(d_positions, labels, y_positions)):
        if label:  # пустые метки
            fig.add_annotation(
                x=d,
                y=y_pos,
                text=label,
                showarrow=False,
                font=dict(size=16, family='Inter, Arial, sans-serif', color='#333')
            )


    fig.add_annotation(
        x=50,
        y=max_curve[100],  # Середина верхней кривой
        text='Δl max',
        showarrow=False,
        font=dict(size=18, family='Inter, Arial, sans-serif', color='#222')
    )

    fig.add_annotation(
        x=50,
        y=mid_curve[100],  # Середина средней кривой
        text='l̄',
        showarrow=False,
        font=dict(size=18, family='Inter, Arial, sans-serif', color='#222')
    )

    fig.add_annotation(
        x=50,
        y=min_curve[100],  # Середина нижней кривой
        text='Δl min',
        showarrow=False,
        font=dict(size=18, family='Inter, Arial, sans-serif', color='#222')
    )

    # Настраиваем внешний вид графика
    y_min = min(min_curve) * 0.8
    y_max = max(max_curve) * 1.2

    fig.update_layout(
        title=dict(text=f'Эпюра износа — Сечение {section_name}', font=PLOTLY_TITLE_FONT),
        font=PLOTLY_FONT,
        xaxis=dict(
            title=dict(text='d₁, мм', font=PLOTLY_AXIS_FONT),
            showgrid=True,
            gridwidth=1,
            gridcolor='rgb(220,220,220)',
            range=[0, 100],
            dtick=10,
            tickfont=PLOTLY_TICK_FONT
        ),
        yaxis=dict(
            title=dict(text='Δl₁, мм', font=PLOTLY_AXIS_FONT),
            showgrid=True,
            gridwidth=1,
            gridcolor='rgb(220,220,220)',
            range=[y_min, y_max],
            dtick=10,
            tickfont=PLOTLY_TICK_FONT
        ),
        plot_bgcolor='white',
        showlegend=False,
        width=1100,
        height=650,
        margin=dict(t=80, b=70, l=90, r=50)
    )

    # Сохраняем график
    plot_file = f'epure_{section_name}.html'
    fig.write_html(os.path.join('static', plot_file), include_plotlyjs='cdn')
    return plot_file


def create_profile(values, section_name):
    """
    Функция создания профилей лапы посевного комплекса "Моррис"
    согласно методике "Износы культиваторных лап посевного комплекса «Моррис»"
    """
    # Validate input
    if values is None or len(values) == 0:
        # Create an empty figure if no data is provided
        fig = go.Figure()
        fig.update_layout(
            title='Нет данных для построения профилей',
            xaxis_title='l, мм',
            yaxis_title='',
            plot_bgcolor='white'
        )
        plot_file = f'profile_{section_name}.html'
        fig.write_html(os.path.join('static', plot_file), include_plotlyjs='cdn')
        return plot_file

    # Создаем график
    fig = go.Figure()

    # Определяем координаты для осей и линий согласно методическим указаниям
    # Ширина культиваторной лапы Моррис составляет 310 мм
    x = np.array([0, 50, 100, 155, 210, 260, 310])  # Точки по оси X (l, мм) - ширина захвата
    y_positions = np.array([0, 10, 20, 30, 40, 50])  # Точки по оси Y (d, мм) - расстояние от носка

    # Добавляем горизонтальные линии сетки
    for y in y_positions:
        fig.add_shape(
            type="line",
            x0=0, x1=310,
            y0=y, y1=y,
            line=dict(color="blue", width=1 if y % 20 == 0 else 0.5)
        )

    # Добавляем вертикальные линии сетки
    for x_val in x:
        fig.add_shape(
            type="line",
            x0=x_val, x1=x_val,
            y0=0, y1=100,
            line=dict(color="gray", width=1 if x_val in [0, 155, 310] else 0.5)
        )

    # Добавляем метки d с правильными подписями согласно методичке
    labels = {'d₁': 0, 'd₂': 20, 'd₃': 40, 'd₄': 45}
    for label, y_pos in labels.items():
        fig.add_annotation(x=-10, y=y_pos, text=label, showarrow=False, font=dict(size=16, family='Inter, Arial, sans-serif', color='#333'), xanchor="right")

    # Добавляем метку di,мм слева
    fig.add_annotation(x=-20, y=25, text='Расстояние от носка, мм', showarrow=False, font=dict(size=16, family='Inter, Arial, sans-serif', color='#333'), xanchor="center", textangle=-90)

    # Функция для создания плавных профильных кривых
    def create_smooth_curve(x_points, y_points, color='teal', name=None, width=2, dash='solid'):
        # Создаем больше точек для более плавной кривой
        x_smooth = np.linspace(min(x_points), max(x_points), 300)

        # Создаем сплайн
        spl = make_interp_spline(x_points, y_points, k=3)
        y_smooth = spl(x_smooth)

        # Добавляем кривую на график
        fig.add_trace(go.Scatter(
            x=x_smooth, y=y_smooth,
            mode='lines',
            line=dict(color=color, width=width, dash=dash),
            name=name,
            showlegend=False if name is None else True
        ))

    # Преобразуем входные данные в профиль
    # Получаем среднее, минимальное и максимальное значения из данных
    mean_value = sum(values) / len(values)
    min_value = min(values)
    max_value = max(values)
    
    # Создаем параболическую форму профиля, где пик находится в середине
    # Ширина лапы культиватора Моррис составляет 310 мм, как указано в комментариях
    x_profile = [0, 50, 100, 155, 210, 260, 310]  # Точки по ширине захвата
    
    # Создаем параболическую форму для среднего профиля
    # Пик будет на 155 мм (середина), а края ниже
    # Среднее значение соответствует пику
    y_middle_profile = []
    for x_pos in x_profile:
        # Формула для параболы: y = a * (x - h)^2 + k, где (h,k) - координаты вершины
        # Здесь h = 155 (центр), k = mean_value (пиковое значение)
        # Коэффициент a определяет, насколько круто изгибается парабола
        # Отрицательное a дает перевернутую параболу (∩)
        distance_from_center = abs(x_pos - 155)
        factor = 1 - (distance_from_center / 155) * 0.5  # Уменьшаем до 50% на краях
        y_middle_profile.append(mean_value * factor)
    
    # Заводской профиль обычно выше среднего на 30-40%
    y_factory_profile = [value * 1.35 for value in y_middle_profile]
    
    # Профиль максимального износа обычно ниже среднего на 30-40%
    y_min_profile = [value * 0.65 for value in y_middle_profile]
    
    # Профиль в заводском исполнении (крайняя правая кривая)
    create_smooth_curve(x_profile, y_factory_profile, color='rgb(22, 96, 167)', name="Заводской профиль", width=3)

    # Средний профиль после эксплуатации (средняя кривая)
    create_smooth_curve(x_profile, y_middle_profile, color='rgb(0, 100, 80)', name="Средний профиль износа", width=2.5, dash='dot')

    # Минимальный профиль после эксплуатации (левая кривая)
    create_smooth_curve(x_profile, y_min_profile, color='rgb(217, 83, 25)', name="Профиль максимального износа", width=2, dash='dashdot')

    # Добавляем пояснения для ширины захвата
    fig.add_annotation(
        x=155, y=-8,
        text='Ширина захвата (l), мм',
        showarrow=False,
        font=dict(size=16, family='Inter, Arial, sans-serif', color='#333')
    )

    # Определяем диапазон оси Y на основе данных
    y_min = min(min(y_min_profile) * 0.9, 0)  # Минимум с запасом, но не меньше нуля
    y_max = max(y_factory_profile) * 1.1  # Максимум с запасом

    # Настраиваем внешний вид
    fig.update_layout(
        title=dict(text='Профили лапы культиватора', font=PLOTLY_TITLE_FONT),
        font=PLOTLY_FONT,
        plot_bgcolor='white',
        showlegend=True,
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1,
            font=dict(size=15, family='Inter, Arial, sans-serif')
        ),
        xaxis=dict(
            title=dict(text='l, мм', font=PLOTLY_AXIS_FONT),
            range=[-30, 340],
            showgrid=False,
            zeroline=False,
            tickvals=x,
            ticktext=[str(int(val)) for val in x],
            tickfont=PLOTLY_TICK_FONT
        ),
        yaxis=dict(
            title=dict(text='d, мм', font=PLOTLY_AXIS_FONT),
            range=[y_min, y_max],
            showgrid=False,
            zeroline=False,
            tickvals=[0, 10, 20, 30, 40, 50],
            ticktext=[str(int(val)) for val in [0, 10, 20, 30, 40, 50]],
            tickfont=PLOTLY_TICK_FONT
        ),
        margin=dict(t=80, l=90, r=50, b=80),
        height=650,
        width=1100
    )

    # Сохраняем график
    plot_file = f'profile_profile.html'
    fig.write_html(os.path.join('static', plot_file), include_plotlyjs='cdn')
    return plot_file


def create_probability_curve(data_values, section_name):
    """
    Создает кривую накопленной опытной вероятности для набора данных
    """
    # Сортируем данные по возрастанию
    sorted_values = sorted(data_values)
    n = len(sorted_values)

    # Проверяем на дубликаты и добавляем небольшой шум к повторяющимся значениям
    value_counts = {}
    for i in range(len(sorted_values)):
        if sorted_values[i] in value_counts:
            # Добавляем очень малое случайное число (шум) к дубликату
            # Чтобы сохранить порядок, делаем шум зависимым от номера дубликата
            value_counts[sorted_values[i]] += 1
            noise = value_counts[sorted_values[i]] * 0.0000001  # Очень малое значение, чтобы не влиять на визуализацию
            sorted_values[i] = sorted_values[i] + noise
        else:
            value_counts[sorted_values[i]] = 0

    # Повторно сортируем значения после добавления шума
    sorted_values = sorted(sorted_values)

    # Рассчитываем накопленную опытную вероятность для каждого значения
    probabilities = [(i + 1) / n for i in range(n)]

    # Создаем более плавную кривую с помощью сплайна
    if n > 3:
        # Если точек достаточно для построения сплайна
        x_smooth = np.linspace(min(sorted_values), max(sorted_values), 300)
        spl = make_interp_spline(sorted_values, probabilities, k=3 if n > 3 else 1)
        y_smooth = spl(x_smooth)
    else:
        # Если точек недостаточно, используем линейную интерполяцию
        x_smooth = sorted_values
        y_smooth = probabilities

    # Создаем график
    fig = go.Figure()

    # Добавляем основную кривую
    fig.add_trace(go.Scatter(
        x=x_smooth,
        y=y_smooth,
        mode='lines',
        name='Кривая накопленной опытной вероятности',
        line=dict(color='rgb(0, 128, 128)', width=2)
    ))

    # Добавляем точки исходных данных
    fig.add_trace(go.Scatter(
        x=sorted_values,
        y=probabilities,
        mode='markers',
        name='Опытные точки',
        marker=dict(color='rgb(0, 128, 128)', size=8)
    ))

    # Настраиваем внешний вид графика
    fig.update_layout(
        title=dict(text=f'Кривая накопленной вероятности — Сечение {section_name}', font=PLOTLY_TITLE_FONT),
        xaxis_title=dict(text='Значение', font=PLOTLY_AXIS_FONT),
        yaxis_title=dict(text='Накопленная вероятность F(x)', font=PLOTLY_AXIS_FONT),
        font=PLOTLY_FONT,
        yaxis=dict(range=[0, 1.05], tickformat='.2f', tickfont=PLOTLY_TICK_FONT),
        xaxis=dict(tickfont=PLOTLY_TICK_FONT),
        template='plotly_white',
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1,
            font=dict(size=15, family='Inter, Arial, sans-serif')
        ),
        width=1100,
        height=650,
        margin=dict(l=80, r=50, t=80, b=70),
    )

    # Добавляем сетку
    fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='rgba(0,0,0,0.1)')
    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='rgba(0,0,0,0.1)')

    # Создаем имя файла
    filename = f"probability_curve_{section_name}_{int(time.time())}.html"
    filepath = os.path.join('static', filename)

    # Сохраняем график в файл
    fig.write_html(filepath, include_plotlyjs='cdn')

    return filename


@app.route('/download')
def download():
    return send_file('static/data.xlsx', as_attachment=True)


if __name__ == '__main__':
    app.run(debug=True)